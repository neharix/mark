import json

from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from .containers import ProjectMarkContainer
from .models import *


def get_projects_schedule(project: Project):
    for schedule in Schedule.objects.all():
        if project.pk in json.loads(schedule.quene_json):
            return schedule
    return None


def get_unparticipated_juries(project: Project, schedule: Schedule):
    juries = []
    for jury in schedule.juries.all():
        if not Mark.objects.filter(jury=jury, project=project).exists():
            juries.append(jury)
    return juries


def numerate_containers(list_of_objects):
    temp_list = []
    index = 1
    for obj in list_of_objects:
        obj.number = index
        temp_list.append(obj)
        index += 1
    return temp_list


def get_juries_marks(jury: User, direction: Direction = None):
    if direction != None:
        return Mark.objects.filter(jury=jury, project__direction=direction).order_by(
            "-date"
        )
    else:
        return Mark.objects.filter(jury=jury).order_by("-date")


def create_worksheet(workbook, direction: Direction = None):
    PRE_COL_SIZE = 2
    POST_COL_SIZE = 3

    font = Font(name="Times New Roman", size=11, color="000000")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    juries = User.objects.filter(groups__name="Jury")
    worksheet: Worksheet = (
        workbook.create_sheet("Umumy", 0)
        if direction == None
        else workbook.create_sheet(direction.name)
    )
    worksheet.column_dimensions["A"].width = 4
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=PRE_COL_SIZE + juries.count() + POST_COL_SIZE,
    )
    header = worksheet.cell(
        1,
        1,
        (
            '"Sanly çözgüt-2024" atly basleşigiň ähli ugurlary boýunça dalaşgärleriniň sanawy'
            if direction == None
            else f'"Sanly çözgüt-2024" atly basleşigiň "{direction.name}" ugry boýunça dalaşgärleriniň sanawy'
        ),  # fix automatic year detection
    )
    header.alignment = Alignment(
        horizontal="center",
    )
    header.font = font

    worksheet.cell(2, 1).border = Border(left=Side("thin"), top=Side("thin"))
    worksheet.cell(3, 1).border = Border(left=Side("thin"))

    worksheet.merge_cells(start_column=2, start_row=2, end_column=2, end_row=3)
    cell = worksheet.cell(2, 2, "Emin agzalary")
    worksheet.cell(3, 2).border = border

    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = font
    cell.border = border
    worksheet.column_dimensions["B"].width = 50
    worksheet.row_dimensions[3].height = 100
    worksheet.merge_cells(
        start_column=PRE_COL_SIZE + juries.count() + 1,
        start_row=2,
        end_column=PRE_COL_SIZE + juries.count() + 1,
        end_row=3,
    )
    cell = worksheet.cell(
        2, PRE_COL_SIZE + juries.count() + 1, "Emin agzalarynyň gatnaşygy"
    )
    cell.alignment = Alignment("center", "center", textRotation=90, wrap_text=True)
    cell.font = font
    cell.border = border
    worksheet.cell(3, PRE_COL_SIZE + juries.count() + 1).border = border

    worksheet.merge_cells(
        start_column=PRE_COL_SIZE + juries.count() + 2,
        start_row=2,
        end_column=PRE_COL_SIZE + juries.count() + 2,
        end_row=3,
    )
    cell = worksheet.cell(
        2,
        PRE_COL_SIZE + juries.count() + 2,
        "Emin agzalary tarapyndan goýulan bahalaryň jemi",
    )
    cell.alignment = Alignment("center", "center", textRotation=90, wrap_text=True)
    cell.font = font
    cell.border = border
    worksheet.cell(3, PRE_COL_SIZE + juries.count() + 2).border = border

    worksheet.merge_cells(
        start_column=PRE_COL_SIZE + juries.count() + 3,
        start_row=2,
        end_column=PRE_COL_SIZE + juries.count() + 3,
        end_row=3,
    )
    cell = worksheet.cell(
        2,
        PRE_COL_SIZE + juries.count() + 3,
        "Ballaryň jemi",
    )
    cell.alignment = Alignment("center", "center", textRotation=90, wrap_text=True)
    cell.font = font
    cell.border = border
    worksheet.cell(3, PRE_COL_SIZE + juries.count() + 3).border = border

    jury_positions = {}
    column_index = 3
    index = 1
    for jury in juries:
        cell = worksheet.cell(2, column_index, index)
        cell.alignment = Alignment("center")
        cell.font = font
        cell.border = border
        cell = worksheet.cell(3, column_index, f"{jury.last_name} {jury.first_name}")
        cell.alignment = Alignment("center", "center", textRotation=90, wrap_text=True)
        cell.font = font
        cell.border = border
        jury_positions[f"{jury.last_name} {jury.first_name}"] = column_index
        column_index += 1
        index += 1

    cell = worksheet.cell(
        4,
        1,
        "T/b",
    )
    cell.alignment = Alignment(
        horizontal="center",
    )
    cell.font = font
    cell.border = border

    cell = worksheet.cell(
        4,
        2,
        "Dalaşgärler",
    )
    cell.alignment = Alignment(
        horizontal="center",
    )
    cell.font = font
    cell.border = border

    worksheet.merge_cells(
        start_row=4,
        start_column=PRE_COL_SIZE + 1,
        end_row=4,
        end_column=PRE_COL_SIZE + juries.count() + POST_COL_SIZE,
    )

    worksheet.cell(4, PRE_COL_SIZE + juries.count() + POST_COL_SIZE).border = border

    cell = worksheet.cell(
        4,
        PRE_COL_SIZE + 1,
        "Emin agzalary tarapyndan, taslamalara goýulan bahalar",
    )
    cell.alignment = Alignment(
        horizontal="center",
    )
    cell.font = font
    cell.border = border

    rated_projects = []
    projects = (
        Project.objects.all()
        if direction == None
        else Project.objects.filter(direction=direction)
    )
    for project in projects:
        if Mark.objects.filter(project=project).exists():
            rated_projects.append(ProjectMarkContainer(project))
    rated_projects.sort(key=lambda e: e.percent)
    rated_projects.reverse()

    row_index = 5
    project_index = 1
    for project in rated_projects:
        cell = worksheet.cell(row_index, 1, project_index)
        cell.alignment = Alignment("center")
        cell.font = font
        cell.border = border

        cell = worksheet.cell(
            row_index, 2, f"{project.agency}\nLideri:{project.manager}"
        )
        cell.alignment = Alignment("center", wrapText=True, wrap_text=True)
        cell.font = font
        cell.border = border
        for mark in project.mark_objects_list:
            cell = worksheet.cell(
                row_index,
                jury_positions[f"{mark.jury.last_name} {mark.jury.first_name}"],
                mark.mark,
            )
            cell.alignment = Alignment("center")
            cell.font = font
            cell.fill = PatternFill(
                start_color="86b8ee", end_color="86b8ee", fill_type="solid"
            )

        for row in worksheet[
            f"{worksheet.cell(row_index, PRE_COL_SIZE + 1).coordinate}:{worksheet.cell(row_index, PRE_COL_SIZE + juries.count()).coordinate}"
        ]:
            for cell in row:
                cell.border = border

        cell = worksheet.cell(
            row_index,
            PRE_COL_SIZE + juries.count() + 1,
            get_projects_schedule(Project.objects.get(pk=project.pk)).juries.count(),
        )
        cell.alignment = Alignment("center")
        cell.font = font
        cell.border = border

        cell = worksheet.cell(
            row_index,
            PRE_COL_SIZE + juries.count() + 2,
            sum([mark_obj.mark for mark_obj in project.mark_objects_list]),
        )
        cell.alignment = Alignment("center")
        cell.font = font
        cell.border = border

        cell = worksheet.cell(
            row_index,
            PRE_COL_SIZE + juries.count() + 3,
            project.percent,
        )
        cell.alignment = Alignment("center")
        cell.font = font
        cell.fill = PatternFill(
            start_color="6fe267", end_color="6fe267", fill_type="solid"
        )
        cell.border = border

        row_index += 1
        project_index += 1

    row_index += 5
    if direction == None:
        cell = worksheet.cell(row_index, 2, "Dalaşgärler")
        cell.alignment = Alignment("center")
        cell.font = font
        cell.border = border
        worksheet.merge_cells(
            start_row=row_index, start_column=3, end_row=row_index, end_column=4
        )
        cell = worksheet.cell(row_index, 3, "Orunlar")
        cell.border = border
        cell.font = font
        cell.alignment = Alignment("center")
        worksheet.cell(row_index, 4).border = border
        worksheet.merge_cells(
            start_row=row_index, start_column=5, end_row=row_index, end_column=7
        )
        cell = worksheet.cell(row_index, 5, "Taslamalar")
        cell.border = border
        cell.font = font
        cell.alignment = Alignment("center")
        worksheet.cell(row_index, 6).border = border
        worksheet.cell(row_index, 7).border = border
        cell = worksheet.cell(row_index, 8, "Jemi")
        cell.border = border
        cell.font = font
        cell.alignment = Alignment("center")
        for direction in Direction.objects.all():
            row_index += 1
            worksheet.merge_cells(
                start_row=row_index, start_column=2, end_row=row_index, end_column=8
            )
            for i in range(2, 9):
                worksheet.cell(row_index, i).border = border
            cell = worksheet.cell(row_index, 2, direction.name)
            cell.font = font
            cell.alignment = Alignment("center")
            for i in range(1, 7):
                cell = worksheet.cell(row_index + i, 2)
                cell.alignment = Alignment("center")
                cell.font = font
                cell.border = border
                worksheet.merge_cells(
                    start_row=row_index + i,
                    start_column=3,
                    end_row=row_index + i,
                    end_column=4,
                )
                cell = worksheet.cell(row_index + i, 3)
                cell.border = border
                cell.font = font
                cell.alignment = Alignment("center")
                worksheet.cell(row_index + i, 4).border = border
                worksheet.merge_cells(
                    start_row=row_index + i,
                    start_column=5,
                    end_row=row_index + i,
                    end_column=7,
                )
                cell = worksheet.cell(row_index + i, 5)
                cell.border = border
                cell.font = font
                cell.alignment = Alignment("center")
                worksheet.cell(row_index + i, 6).border = border
                worksheet.cell(row_index + i, 7).border = border
                cell = worksheet.cell(row_index + i, 8)
                cell.border = border
                cell.font = font
                cell.alignment = Alignment("center")
                match i:
                    case 1:
                        worksheet.cell(row_index + i, 3, "1-nji orun")
                        cell.fill = PatternFill(
                            start_color="e0e745", end_color="e0e745", fill_type="solid"
                        )
                    case 2 | 3:
                        worksheet.cell(row_index + i, 3, "2-nji orun")
                        cell.fill = PatternFill(
                            start_color="c3c3c3", end_color="c3c3c3", fill_type="solid"
                        )
                    case 4 | 5 | 6:
                        worksheet.cell(row_index + i, 3, "3-nji orun")
                        cell.fill = PatternFill(
                            start_color="e98c38", end_color="e98c38", fill_type="solid"
                        )

            r_projects = []
            for project in Project.objects.filter(direction=direction):
                if Mark.objects.filter(project=project).exists():
                    r_projects.append(ProjectMarkContainer(project))
            r_projects.sort(key=lambda e: e.percent)
            r_projects.reverse()
            r_projects = r_projects[:6]
            for i in range(1, 7):
                try:
                    worksheet.cell(row_index + i, 2, r_projects[i - 1].manager)
                    worksheet.cell(row_index + i, 5, r_projects[i - 1].name)
                    worksheet.cell(row_index + i, 8, r_projects[i - 1].percent)
                except:
                    pass
            row_index += 6
    else:
        cell = worksheet.cell(row_index, 2, "Dalaşgärler")
        cell.alignment = Alignment("center")
        cell.font = font
        cell.border = border
        worksheet.merge_cells(
            start_row=row_index, start_column=3, end_row=row_index, end_column=4
        )
        cell = worksheet.cell(row_index, 3, "Orunlar")
        cell.border = border
        cell.font = font
        cell.alignment = Alignment("center")
        worksheet.cell(row_index, 4).border = border
        worksheet.merge_cells(
            start_row=row_index, start_column=5, end_row=row_index, end_column=7
        )
        cell = worksheet.cell(row_index, 5, "Taslamalar")
        cell.border = border
        cell.font = font
        cell.alignment = Alignment("center")
        worksheet.cell(row_index, 6).border = border
        worksheet.cell(row_index, 7).border = border
        cell = worksheet.cell(row_index, 8, "Jemi")
        cell.border = border
        cell.font = font
        cell.alignment = Alignment("center")
        for i in range(1, 7):
            cell = worksheet.cell(row_index + i, 2)
            cell.alignment = Alignment("center")
            cell.font = font
            cell.border = border
            worksheet.merge_cells(
                start_row=row_index + i,
                start_column=3,
                end_row=row_index + i,
                end_column=4,
            )
            cell = worksheet.cell(row_index + i, 3)
            cell.border = border
            cell.font = font
            cell.alignment = Alignment("center")
            worksheet.cell(row_index + i, 4).border = border
            worksheet.merge_cells(
                start_row=row_index + i,
                start_column=5,
                end_row=row_index + i,
                end_column=7,
            )
            cell = worksheet.cell(row_index + i, 5)
            cell.border = border
            cell.font = font
            cell.alignment = Alignment("center")
            worksheet.cell(row_index + i, 6).border = border
            worksheet.cell(row_index + i, 7).border = border
            cell = worksheet.cell(row_index + i, 8)
            cell.border = border
            cell.font = font
            cell.alignment = Alignment("center")
            match i:
                case 1:
                    worksheet.cell(row_index + i, 3, "1-nji orun")
                    cell.fill = PatternFill(
                        start_color="e0e745", end_color="e0e745", fill_type="solid"
                    )
                case 2 | 3:
                    worksheet.cell(row_index + i, 3, "2-nji orun")
                    cell.fill = PatternFill(
                        start_color="c3c3c3", end_color="c3c3c3", fill_type="solid"
                    )
                case 4 | 5 | 6:
                    worksheet.cell(row_index + i, 3, "3-nji orun")
                    cell.fill = PatternFill(
                        start_color="e98c38", end_color="e98c38", fill_type="solid"
                    )

        r_projects = rated_projects[:6]
        for i in range(1, 7):
            try:
                worksheet.cell(
                    row_index + i, 2, r_projects[i - 1].manager
                ).alignment = Alignment("center", wrapText=True)
                worksheet.cell(row_index + i, 5, r_projects[i - 1].name).alignment = (
                    Alignment("center", wrapText=True)
                )
                worksheet.cell(row_index + i, 8, r_projects[i - 1].percent)
            except:
                pass
        row_index += 6

    return workbook
